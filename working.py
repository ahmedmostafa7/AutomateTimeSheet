from openpyxl import Workbook, load_workbook
from datetime import date

import os.path

from openpyxl.worksheet import worksheet
import kivy
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from openpyxl.utils import get_column_letter
from kivy.uix.popup import Popup
# from kivy.garden.circulardatetimepicker import CircularTimePicker

from openpyxl.styles import Font


def show_popup():
    pass


class MyGrid(GridLayout):
    def __init__(self, **kwargs):
        super(MyGrid, self).__init__(**kwargs)
        self.cols = 2
        self.add_widget(Label(text="From: "))
        self.fromDate = TextInput(multiline=False)
        self.add_widget(self.fromDate)

        self.add_widget(Label(text="To: "))
        self.toDate = TextInput(multiline=False)
        self.add_widget(self.toDate)

        self.add_widget(Label(text="Task: "))
        self.task = TextInput(multiline=False)
        self.add_widget(self.task)

        self.add_widget(Label(text="Month: "))
        self.month = TextInput(multiline=False)
        self.add_widget(self.month)

        self.add_widget(Label(text="date: "))
        self.date = TextInput(multiline=False)
        self.add_widget(self.date)

        self.sumbit = Button(text="Sumbit", font_size=40)
        self.sumbit.bind(on_press=self.pressed)
        self.add_widget(self.sumbit)

    def pressed(self, instance):
        print("GOt It")

        fromDate = int(self.fromDate.text)
        toDate = int(self.toDate.text)
        task = self.task.text
        month = self.month.text or '26_x_2021'

        if int(fromDate) < 12:
            fromDate = fromDate
        else:
            fromDate = fromDate+12

        print(fromDate, toDate)

        totalHours = ((toDate+12)-fromDate)

        date = self.date.text or "sheet_1"

        # sheetNames=
        listDats = [fromDate, toDate, task, totalHours,  date, month]
        jsonData = {"fromDate": fromDate, "toDate": toDate,
                    "task": task, "totalHours": totalHours, "date": date}
        # arr.append(listDats)
        self.WorkSheet(listDats)

    def WorkSheet(self, data):

        filename = f"timeSheet_{data[5]}.xlsx"
        print(filename)
        if os.path.isfile(filename):

            wb = load_workbook(filename)
            newData = data[:4]
            print(wb.sheetnames, wb)
            ws = wb.active
            ws.append(["From", "To", "Task", "Hours"])
            # index = wb.sheetnames.index(data[4])
            if data[4] in wb.sheetnames:
                ws = wb[data[4]]

                ws.append(newData)
            else:
                ws = wb.create_sheet(data[4])
                ws.append(["From", "To", "Task", "Hours"])
                ws.title = data[4]
                ws.append(newData)

            for gela in range(1, 5):

                ws[get_column_letter(gela)+"1"].font = Font(bold=True)
                print(get_column_letter(gela)+"1")

            wb.save(filename)
        else:
            wb = Workbook()
            ws = wb.create_sheet(data[4])
            ws = wb.active
            ws.title = data[4]
            newData = data[:4]
            for col in range(1, 2):
                ws.append(["From", "To", "Task", "Hours"])

                ws.append(newData)
            for gela in range(1, 5):
                ws[get_column_letter(gela)+"1"].font = Font(bold=True)
                print(get_column_letter(gela)+"1")
            wb.save(filename=filename)

        # fileMonth = xlsxwriter.Workbook(filename)


class MyApp(App):

    def build(self):
        return MyGrid()


if __name__ == "__main__":
    MyApp().run()
